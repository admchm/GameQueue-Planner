from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class ExcelStyles(object):
    def __init__(self):
        
        # alignment
        self._cell_content_indent = Alignment(indent=1, horizontal='left', vertical='top')
        
        # sides
        self._dark_grey_side = Side(style='thin', color="b5b8b7")
        self._black_side = Side(style='thin', color="000000")
        
        # borders
        self._first_column_border = Border(left=self._dark_grey_side, right=self._black_side, top=self._dark_grey_side, bottom=self._dark_grey_side)
        
        # fill colors
        self._lighter_grey_fill = PatternFill(start_color="cfcfcf", end_color="cfcfcf", fill_type="solid")
        self._darker_grey_fill = PatternFill(start_color="b5b8b7", end_color="b5b8b7", fill_type="solid")
        
        # fonts
        self._standard_font = Font(name='Calibri', size=12)
        self._standard_bold_font = Font(name='Calibri', size=12, bold=True)
        
    def get_cell_content_indent(self):
        return self._cell_content_indent
    
    def get_first_column_border(self):
        return self._first_column_border
    
    def get_lighter_grey_fill(self):
        return self._lighter_grey_fill
    
    def get_darker_grey_fill(self):
        return self._darker_grey_fill
    
    def get_standard_font(self):
        return self._standard_font
    
    def get_standard_bold_font(self):
        return self._standard_bold_font
    
    def get_black_side(self):
        return self.black_side
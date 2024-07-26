import pandas as pd

from datetime import datetime
from collections import defaultdict, OrderedDict
from openpyxl import load_workbook

from models.Constants.ConstRes import ConstRes
from models.Files.Excel.ExcelStyles import ExcelStyles
from common.LoggerSingleton import LoggerSingleton

class ExcelFileCreator(object):
    def __init__(self, path_combined = ''):
        self.path_combined = path_combined
        self.logger = LoggerSingleton()
        
    def process_file(self, games_list, file_path):
        self.logger.log_info("Preparing Excel file")
        
        self.path_combined = file_path
        
        splitted_games = self.split_games_from_list_to_sheets_per_years(games_list)
        prepared_list = self.sort_games_list_with_years(splitted_games)
        
        data_frames = {}
        for year, game_objects in prepared_list.items():
            data_frames[year] = self.create_dataframe(game_objects)
        
        with pd.ExcelWriter(self.path_combined, engine=ConstRes.OPENPYXL.value) as writer:
            for year, df in data_frames.items():
                df.to_excel(writer, sheet_name=str(year), index=False)
        
        self.adjust_column_widths()
        
        self.logger.log_info(f"Created file at: {self.path_combined}")
        print(f"Created file at: {self.path_combined}")
    
    def split_games_from_list_to_sheets_per_years(self, games_list):
        splitted_games_by_year = defaultdict(list)
        
        for game in games_list:
            if isinstance(game.first_release_date, str):
                # release_year - key in dict
                release_year = datetime.strptime(game.first_release_date, "%Y-%m-%d").year
                splitted_games_by_year[release_year].append(game)
            else:
                self.logger.log_error(f"ERROR: Wrong date format of {game.title}")
        
        return splitted_games_by_year
               
    def sort_games_list_with_years(self, splitted_games):    
        sorted_keys = sorted(splitted_games.keys())
        
        return OrderedDict((key, splitted_games[key]) for key in sorted_keys)
                
    def create_dataframe(self, game_objects):
        data = {
            "Title": [obj.title for obj in game_objects],
            "Platform name": [obj.platform_name for obj in game_objects],
            "Release date": [obj.first_release_date for obj in game_objects],
            "Score": [f"{obj.moby_score} ({obj.moby_num_votes})" for obj in game_objects]
        }
        return pd.DataFrame(data)
    
    def adjust_column_widths(self):
        styles = ExcelStyles()
        wb = load_workbook(self.path_combined)
        
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                    cell.font = styles.get_standard_font()
                    if cell.row == 1:
                        cell.font = styles.get_standard_bold_font()
                        cell.fill = styles.get_darker_grey_fill()
                        
                    if cell.row > 1:
                        cell.alignment = styles.get_cell_content_indent()
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width
            
            for cell in ws['A']: # bold the first column, fill with color
                if cell.row != 1:
                    cell.font = styles.get_standard_bold_font()
                    cell.fill = styles.get_lighter_grey_fill()
                    cell.border = styles.get_first_column_border()
                    
        # Set the zoom level for the worksheet
            ws.sheet_view.zoomScale = 120
        
        wb.save(self.path_combined)
    
                    
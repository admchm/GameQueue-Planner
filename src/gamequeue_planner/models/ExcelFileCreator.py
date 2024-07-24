import pandas as pd
import os

from datetime import datetime
from collections import defaultdict, OrderedDict
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models.DatesEditor import DatesEditor

class ExcelFileCreator:
    
    def __init__(self, path = "~/", file_name = "GameQueue.xlsx", path_combined = ""):
        self.path = path
        self.file_name = file_name
        self.path_combined = path_combined
    
    def set_path(self, path):
        self.path = path
        
    def set_file_name(self, file_name):
        self.file_name = file_name
        
    def add_date_to_file_name(self):
        index = self.file_name.find('.') # inserting before .xlsx
        time = DatesEditor.get_current_time(self)
        
        if index == -1:
            return self.file_name
        
        self.file_name = self.file_name[:index] + time + self.file_name[index:]
                
    def combine_path_with_file_name(self):
        self.path_combined = os.path.expanduser(f"{self.path + self.file_name}")
        
    def split_games_from_list_to_sheets_per_years(self, games_list):
        splitted_games_by_year = defaultdict(list)
        
        for game in games_list:
            if isinstance(game.first_release_date, str):
                # release_year - key in dict
                release_year = datetime.strptime(game.first_release_date, "%Y-%m-%d").year
                splitted_games_by_year[release_year].append(game)
            else:
                print(f"ERROR: Wrong date format of {game.title}")
        
        return splitted_games_by_year
               
    def sort_games_list_with_years(self, splitted_games):    
        sorted_keys = sorted(splitted_games.keys())
        
        return OrderedDict((key, splitted_games[key]) for key in sorted_keys)
                
    def create_dataframe(self, game_objects):
        data = {
            "Title": [obj.title for obj in game_objects],
            "Platform name": [obj.platform_name for obj in game_objects],
            "Release date": [obj.first_release_date for obj in game_objects],
            "Score": [f"{obj.moby_score} ({obj.moby_num_votes})" for obj in game_objects]
        }
        return pd.DataFrame(data)
    
    def adjust_column_widths(self):
        wb = load_workbook(self.path_combined) # self.file_name
        
        lighter_grey_fill = PatternFill(start_color="cfcfcf", end_color="cfcfcf", fill_type="solid")
        darker_grey_fill = PatternFill(start_color="b5b8b7", end_color="b5b8b7", fill_type="solid")
        dark_grey_side = Side(style='thin', color="b5b8b7")
        black_side = Side(style='thin', color="000000")
        
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                    cell.font = Font(name='Calibri', size=12)
                    if cell.row == 1:
                        cell.font = Font(name='Calibri', size=12, bold=True)
                        cell.fill = darker_grey_fill
                        
                    if cell.row > 1:
                        cell.alignment = Alignment(indent=1, horizontal='left', vertical='top')
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width
            
            # bold the first column, fill with color
            for cell in ws['A']:
                if cell.row != 1:
                    cell.font = Font(name='Calibri', size=12, bold=True)
                    cell.fill = lighter_grey_fill
                    cell.border = Border(left=dark_grey_side, right=black_side, top=dark_grey_side, bottom=dark_grey_side)
                    
        # Set the zoom level for the worksheet
            ws.sheet_view.zoomScale = 120
        
        wb.save(self.path_combined)
    
    def prepare_file(self, games_list):
        print("LOG: Preparing Excel file")
        
        #TODO: - Split these functions
        self.add_date_to_file_name()
        self.combine_path_with_file_name()
        splitted_games = self.split_games_from_list_to_sheets_per_years(games_list)
        prepared_list = self.sort_games_list_with_years(splitted_games)
        
        data_frames = {}
        for year, game_objects in prepared_list.items():
            data_frames[year] = self.create_dataframe(game_objects)
        
        with pd.ExcelWriter(self.path_combined, engine='openpyxl') as writer: # required openpyxl
            for year, df in data_frames.items():
                df.to_excel(writer, sheet_name=str(year), index=False)
        
        self.adjust_column_widths()
        
        print(f"LOG: Created file at: {self.path_combined}")                    